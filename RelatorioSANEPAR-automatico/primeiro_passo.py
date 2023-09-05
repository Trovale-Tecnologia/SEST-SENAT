import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import pandas as pd
import win32com.client

# TR_USER
def filter_and_copy(sheet_source, sheet_dest):
    # Filtrar na COLUNA D apenas PJ
    filtered_rows = [row for row in sheet_source.iter_rows(min_row=2, max_col=4, values_only=True) if row[3] == "PJ"]

    # Copiar o CNPJ de todos os registros da coluna A e colar como valores na planilha CNPJ
    for row in filtered_rows:
        sheet_dest.append([row[0]])

# Função para limpar uma coluna da planilha
def clear_column(sheet, column_index):
    if sheet == workbook_ajuste_base['AJUSTE_BASE']:
        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=12):
            for cell in row:
                cell.value = None
    else:
        for row in sheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index):
            for cell in row:
                cell.value = None

# Carregar os arquivos
workbook_cnpj = openpyxl.load_workbook('CNPJ.xlsx')
workbook_ajuste_base = openpyxl.load_workbook('AJUSTE_BASE.xlsx')
workbook_matriz_estudo = openpyxl.load_workbook('Matriz-Estudo-Mensal.xlsx')
workbook_tr_user1 = openpyxl.load_workbook('Relatorio_TR_USER1.xlsx')
workbook_tr_user2 = openpyxl.load_workbook('Relatorio_TR_USER2.xlsx')
workbook_tr_user3 = openpyxl.load_workbook('Relatorio_TR_USER3.xlsx')
workbook_tr_user4 = openpyxl.load_workbook('Relatorio_TR_USER4.xlsx')

# Limpar coluna A da planilha CNPJ
clear_column(workbook_cnpj['CNPJ'], column_index=1)

# Limpar a planilha TODA do Arquivo AJUSTE_BASE
clear_column(workbook_ajuste_base['AJUSTE_BASE'], column_index=1)

# Limpar coluna A da aba BASE.PF do Arquivo Matriz Estudo Mensal
clear_column(workbook_matriz_estudo['BASE.PF'], column_index=1)

# Limpar coluna A da aba BASE.Empresas do Arquivo 2023 - Matriz Estudo Mensal
clear_column(workbook_matriz_estudo['BASE.Empresas'], column_index=1)

# Limpar coluna A e B da aba BASE.Socios do Arquivo 2023 - Matriz Estudo Mensal
clear_column(workbook_matriz_estudo['BASE.SOCIOS'], column_index=1)
clear_column(workbook_matriz_estudo['BASE.SOCIOS'], column_index=2)

# TR_USER 1
filter_and_copy(workbook_tr_user1['Relatório de Consultas'], workbook_cnpj['CNPJ'])

# TR_USER 2
filter_and_copy(workbook_tr_user2['Relatório de Consultas'], workbook_cnpj['CNPJ'])

# TR_USER 3
filter_and_copy(workbook_tr_user3['Relatório de Consultas'], workbook_cnpj['CNPJ'])

# TR_USER 4
filter_and_copy(workbook_tr_user4['Relatório de Consultas'], workbook_cnpj['CNPJ'])


# Remover duplicatas na planilha CNPJ
cnpj_values = set()
for row in workbook_cnpj['CNPJ'].iter_rows(min_row=2, max_col=1, values_only=True):
    cnpj_values.add(row[0])

workbook_cnpj.create_sheet('CNPJfiltrado')
sheet_cnpj_filtrado = workbook_cnpj['CNPJfiltrado']

for cnpj in cnpj_values:
    sheet_cnpj_filtrado.append([cnpj])

# Copiar CNPJs da Base Ajustada (Arquivo CNPJ) para a aba BASE.Empresas no arquivo Matriz-Estudo-Mensal
for row in workbook_cnpj['CNPJfiltrado'].iter_rows(min_row=2, max_col=1, values_only=True):
    workbook_matriz_estudo['BASE.Empresas'].append(row)

# Salvar as modificações nos arquivos
workbook_cnpj.save('CNPJ.xlsx')
workbook_ajuste_base.save('AJUSTE_BASE.xlsx')
workbook_matriz_estudo.save('Matriz-Estudo-Mensal.xlsx')

print('\nProcessamento das planilhas concluída com sucesso!\n')

input('\nPressione Enter para continuar com outros processamentos...')

# Passo 1: Selecionar todas as células da planilha CNPJ_QSA e copiá-las
cnpj_qsa_sheet = workbook_cnpj['CNPJ_QSA']
cnpj_qsa_values = []

for row in cnpj_qsa_sheet.iter_rows(min_row=1, max_row=cnpj_qsa_sheet.max_row, max_col=cnpj_qsa_sheet.max_column, values_only=True):
    cnpj_qsa_values.append(row)


# Passo 2: Abrir a planilha AJUSTE_BASE e colar as células copiadas
ajuste_base_sheet = workbook_ajuste_base['AJUSTE_BASE']

# Inserir as células copiadas no início da planilha AJUSTE_BASE
for row_data in cnpj_qsa_values:
    ajuste_base_sheet.append(row_data)

# Passo 3: Aplicar a lógica de limpeza na planilha AJUSTE_BASE
# Aqui você pode realizar as operações necessárias para ajustar a planilha AJUSTE_BASE
# Por exemplo, se você deseja excluir colunas específicas, pode fazer isso aqui usando o pandas ou openpyxl.
# Suponha que você deseja excluir as colunas B, C e D da planilha AJUSTE_BASE:
columns_to_delete = ['B', 'C', 'D']

for col in columns_to_delete:
    col_index = openpyxl.utils.column_index_from_string(col)
    for row in ajuste_base_sheet.iter_rows(min_row=1, max_row=ajuste_base_sheet.max_row, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.value = None

