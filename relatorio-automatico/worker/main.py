import openpyxl
import pandas as pd

# Função para verificar se um valor pode ser convertido em um número
def is_numeric(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# TR_USER
def filter_and_copy(sheet_source, sheet_dest):
    # Filtrar na COLUNA D apenas PJ
    filtered_rows = [row for row in sheet_source.iter_rows(min_row=2, max_col=4, values_only=True) if row[3] == "PJ"]

    # Copiar o CNPJ de todos os registros da coluna A e colar como valores na planilha CNPJ
    for row in filtered_rows:
        sheet_dest.append([row[0]])

# Função para limpar uma coluna da planilha
def clear_column(sheet, column_index):
    if sheet == workbook_ajuste_base['AJUSTE_BASE']:
        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=12):
            for cell in row:
                cell.value = None
    else:
        for row in sheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index):
            for cell in row:
                cell.value = None

# Carregar os arquivos
workbook_cnpj = openpyxl.load_workbook('CNPJ.xlsx')
workbook_ajuste_base = openpyxl.load_workbook('AJUSTE_BASE.xlsx')
workbook_matriz_estudo = openpyxl.load_workbook('Matriz Estudo Mensal.xlsx')
workbook_tr_user1 = openpyxl.load_workbook('Relatorio_TR_USER1.xlsx')
workbook_tr_user2 = openpyxl.load_workbook('Relatorio_TR_USER2.xlsx')
workbook_tr_user3 = openpyxl.load_workbook('Relatorio_TR_USER3.xlsx')
workbook_tr_user4 = openpyxl.load_workbook('Relatorio_TR_USER4.xlsx')

# Limpar coluna A da planilha CNPJ
clear_column(workbook_cnpj['CNPJ'], column_index=1)

# Limpar a planilha TODA do Arquivo AJUSTE_BASE
clear_column(workbook_ajuste_base['AJUSTE_BASE'], column_index=1)

# Limpar coluna A da aba BASE.PF do Arquivo Matriz Estudo Mensal
clear_column(workbook_matriz_estudo['BASE.PF'], column_index=1)

# Limpar coluna A da aba BASE.Empresas do Arquivo 2023 - Matriz Estudo Mensal
clear_column(workbook_matriz_estudo['BASE.Empresas'], column_index=1)

# Limpar coluna A e B da aba BASE.Socios do Arquivo 2023 - Matriz Estudo Mensal
clear_column(workbook_matriz_estudo['BASE.SOCIOS'], column_index=1)
clear_column(workbook_matriz_estudo['BASE.SOCIOS'], column_index=2)

# TR_USER 1
filter_and_copy(workbook_tr_user1['Relatório de Consultas'], workbook_cnpj['CNPJ'])

# TR_USER 2
filter_and_copy(workbook_tr_user2['Relatório de Consultas'], workbook_cnpj['CNPJ'])

# TR_USER 3
filter_and_copy(workbook_tr_user3['Relatório de Consultas'], workbook_cnpj['CNPJ'])

# TR_USER 4
filter_and_copy(workbook_tr_user4['Relatório de Consultas'], workbook_cnpj['CNPJ'])


# Remover duplicatas na planilha CNPJ
cnpj_values = set()
for row in workbook_cnpj['CNPJ'].iter_rows(min_row=2, max_col=1, values_only=True):
    cnpj_values.add(row[0])

sheet_cnpj_filtrado = workbook_cnpj['CNPJ']

for cnpj in cnpj_values:
    sheet_cnpj_filtrado.append([cnpj])

# Copiar CNPJs da Base Ajustada (Arquivo CNPJ) para a aba BASE.Empresas no arquivo Matriz-Estudo-Mensal
for row in workbook_cnpj['CNPJ'].iter_rows(min_row=2, max_col=1, values_only=True):
    workbook_matriz_estudo['BASE.Empresas'].append(row)

# Salvar as modificações nos arquivos
workbook_cnpj.save('CNPJ.xlsx')
workbook_ajuste_base.save('AJUSTE_BASE.xlsx')
workbook_matriz_estudo.save('Matriz Estudo Mensal.xlsx')

print('\nProcessamento dos CNPJs concluído com sucesso!\n')

# Selecionar todas as células da planilha CNPJ_QSA e colar na PLANILHA AJUSTE_BASE

# Abra o arquivo Excel da planilha de origem (substitua 'CNPJs_QSA.xlsx' pelo nome do seu arquivo)
workbook_origem = openpyxl.load_workbook('CNPJs_QSA.xlsx')

# Abra o arquivo Excel da planilha de destino (substitua 'seuarquivo_destino.xlsx' pelo nome do seu arquivo)
workbook_destino = openpyxl.load_workbook('AJUSTE_BASE.xlsx')

# Selecione a planilha de origem
planilha_origem = workbook_origem['Registros_Datahub']

# Selecione a planilha de destino
planilha_destino = workbook_destino['AJUSTE_BASE']

# Percorra todas as células na planilha de origem e copie seus valores para a planilha de destino
for row in planilha_origem.iter_rows():
    for cell in row:
        # Copie o valor da célula
        valor_celula = cell.value
        
        # Cole o valor na planilha de destino na mesma posição
        planilha_destino[cell.coordinate].value = valor_celula

# Salve as mudanças no arquivo de destino
workbook_destino.save('AJUSTE_BASE.xlsx')

print('\n Processamento de CNPJs feito!\n')

input("\n Precione ENTER uma vez que a planilha CNPJs_QSA for introduzida no diretório local do script...")

print('\n Todas as células da planilha CNPJ_QSA foram coladas na PLANILHA AJUSTE_BASE!\n')


# Parte da MACRO



# Carregue o arquivo Excel (substitua 'seuarquivo.xlsx' pelo nome do seu arquivo)
workbook_ajuste_base = openpyxl.load_workbook('AJUSTE_BASE.xlsx')

# Selecione a planilha desejada (substitua 'Planilha1' pelo nome da sua planilha)
sheet = workbook_ajuste_base['AJUSTE_BASE']

# Lista das colunas a serem excluídas em ordem (começando em 1)
colunas_a_excluir = [
    ('B', 'D'),  # Colunas B a D
    ('C', 'E'),  # Colunas C a E
    ('D', 'G'),  # Colunas D a G
    ('E', 'G'),  # Colunas E a G
    ('F', 'I'),  # Colunas F a I
    ('G', 'I'),  # Colunas G a I
    ('H', 'J'),  # Colunas H a J
]

# Exclua as colunas individualmente, na mesma ordem do código VBA
for coluna_inicio, coluna_fim in colunas_a_excluir:
    coluna_inicio_index = openpyxl.utils.column_index_from_string(coluna_inicio)
    coluna_fim_index = openpyxl.utils.column_index_from_string(coluna_fim)
    
    # Exclua apenas as colunas especificadas, mantendo as outras intactas
    for i in range(coluna_inicio_index, coluna_fim_index + 1):
        sheet.delete_cols(coluna_inicio_index)

# Salve as mudanças no arquivo Excel (ou substitua o original)
workbook_ajuste_base.save('AJUSTE_BASE.xlsx')

# Carregue o arquivo Excel
df = pd.read_excel('AJUSTE_BASE.xlsx', sheet_name='AJUSTE_BASE')

# Selecione a coluna CNPJ
cnpj_column = df['CNPJ']

# Transformar outras colunas em linhas
df = df.melt(id_vars=['CNPJ'], var_name='Atributo', value_name='Valor')

# Remover linhas com valores em branco na coluna "Valor"
df = df.dropna(subset=['Valor'])

# Remover duplicatas na coluna "Valor"
df = df.drop_duplicates(subset=['Valor'])

# Salvar o DataFrame de volta no arquivo Excel
df.to_excel('AJUSTE_BASE.xlsx', sheet_name='AJUSTE_BASE', index=False)



# PARTE FINAL - CÓPIA DE CPFs



# Lista de nomes de arquivo dos relatórios TR_USER
arquivos_tr_user = ['Relatorio_TR_USER1.xlsx', 'Relatorio_TR_USER2.xlsx', 'Relatorio_TR_USER3.xlsx', 'Relatorio_TR_USER4.xlsx']

# Carregar o arquivo Excel AJUSTE_BASE.xlsx
workbook_ajuste_base = openpyxl.load_workbook('AJUSTE_BASE.xlsx')

# Carregar o arquivo Excel Matriz Estudo Mensal.xlsx
workbook_matriz_estudo = openpyxl.load_workbook('Matriz Estudo Mensal.xlsx')

# Selecionar a planilha BASE.SOCIOS
sheet_base_socios = workbook_matriz_estudo['BASE.SOCIOS']

# Loop através dos relatórios TR_USER
for idx, arquivo_tr_user in enumerate(arquivos_tr_user, start=1):
    # Carregar o arquivo TR_USER atual
    workbook_tr_user = openpyxl.load_workbook(arquivo_tr_user)

    # Selecionar a planilha correspondente (substitua pelo nome da planilha)
    sheet_tr_user = workbook_tr_user['Relatório de Consultas']

    # Selecionar a coluna "CPF" na planilha TR_USER
    cpf_column_tr_user = sheet_tr_user['A']

    # Filtrar na COLUNA D apenas PF
    cpf_values_filtered = [cell.value for cell in cpf_column_tr_user if cell.offset(column=3).value == 'PF']

    # Copiar os CPFs filtrados para a planilha BASE.PF na planilha Matriz Estudo Mensal
    sheet_base_pf = workbook_matriz_estudo['BASE.PF']
    for cpf in cpf_values_filtered[0:]:  
        sheet_base_pf.append([cpf])

    # Salvar as mudanças no arquivo TR_USER atual
    workbook_tr_user.save(arquivo_tr_user)

# Passo 1: Copiar os CPFs da Base Ajustada (Arquivo > AJUSTE_BASE - COLUNA C) para a planilha BASE.SOCIOS
cpf_column_ajuste_base = workbook_ajuste_base['AJUSTE_BASE']['C']
cpf_values = [cell.value for cell in cpf_column_ajuste_base]
for cpf in cpf_values[1:]:  # Começando a partir da segunda linha para evitar o cabeçalho
    # Adicionar um critério condicional para evitar valores iguais a 0 na coluna A
    if cpf != 0:
        sheet_base_socios.append([cpf])

# Passo 6: Substituir "." e "-" por VAZIO (CTRL + SHIFT + D) na planilha BASE.PF
sheet_base_pf = workbook_matriz_estudo['BASE.PF']
for row in sheet_base_pf.iter_rows(min_row=2, max_row=sheet_base_pf.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value:
            cell.value = cell.value.replace(".", "").replace("-", "")

# Passo 7: Na planilha BASE.PF remover duplicatas do CPF
def remover_duplicatas(sheet):
    data = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        data.append(row[0])

    # Remover duplicatas
    data = list(set(data))

    # Limpar a planilha
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = None

    # Escrever dados únicos de volta na planilha
    for idx, value in enumerate(data, start=2):
        sheet.cell(row=idx, column=1, value=value)

# Remover duplicatas das planilhas BASE.PF, BASE.SOCIOS e BASE.Empresas
remover_duplicatas(workbook_matriz_estudo['BASE.PF'])
remover_duplicatas(workbook_matriz_estudo['BASE.SOCIOS'])
remover_duplicatas(workbook_matriz_estudo['BASE.Empresas'])

# Salvar as mudanças nos arquivos Excel
workbook_ajuste_base.save('AJUSTE_BASE.xlsx')
workbook_matriz_estudo.save('Matriz Estudo Mensal.xlsx')
